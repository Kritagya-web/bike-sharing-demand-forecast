"""
Daily bike-demand report Lambda: scenario rows -> SageMaker endpoint -> Excel + chart -> SES email.

Environment variables:
  ENDPOINT_NAME (required)
  REPORT_EMAIL_FROM, REPORT_EMAIL_TO (required for SES)
  REPORT_TIMEZONE (default America/New_York): used for "tomorrow" date and email timestamps
  AWS_REGION (default from Lambda)
  TARGET_DATE (optional YYYY-MM-DD): override forecast day for tests
  ASSUME_HOLIDAY (default 0): 1 forces holiday / non-workingday pattern
  ASSUME_YR (default 1): yr feature as in training data (0=first year, 1=second year of dataset)
"""

from __future__ import annotations

import io
import json
import logging
import os
from datetime import date, datetime, timedelta
from email import policy
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import boto3
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger()
logger.setLevel(logging.INFO)

PROFILE_PATH = Path(__file__).resolve().parent / "hourly_profile.json"


def _load_profile() -> dict[str, Any]:
    return json.loads(PROFILE_PATH.read_text(encoding="utf-8"))


def _bike_weekday(d: date) -> int:
    """UCI bike sharing weekday: 0=Sunday .. 6=Saturday."""
    py = d.weekday()  # Monday=0 .. Sunday=6
    return (py + 1) % 7


def build_instances(target: date, profile: dict[str, Any]) -> list[dict[str, Any]]:
    mnth = target.month
    season = int(profile["season_by_month"][str(mnth)])
    bike_wd = _bike_weekday(target)
    holiday = int(os.environ.get("ASSUME_HOLIDAY", "0"))
    workingday = 0 if (holiday or bike_wd in (0, 6)) else 1
    yr = int(os.environ.get("ASSUME_YR", "1"))
    hours = {int(h["hour"]): h for h in profile["hours"]}
    rows: list[dict[str, Any]] = []
    for hr in range(24):
        ph = hours[hr]
        rows.append(
            {
                "instant": hr + 1,
                "dteday": target.isoformat(),
                "season": season,
                "yr": yr,
                "mnth": mnth,
                "hr": hr,
                "holiday": holiday,
                "weekday": bike_wd,
                "workingday": workingday,
                "weathersit": ph["weathersit"],
                "temp": ph["temp"],
                "atemp": ph["atemp"],
                "hum": ph["hum"],
                "windspeed": ph["windspeed"],
            }
        )
    return rows


def _parse_predictions(raw: str) -> list[float]:
    data = json.loads(raw)
    if isinstance(data, dict) and "predictions" in data:
        data = data["predictions"]
    out: list[float] = []
    for x in data:
        if isinstance(x, (list, tuple)):
            out.extend(float(t) for t in x)
        elif isinstance(x, dict):
            v = x.get("score", x.get("prediction"))
            if v is None and x:
                v = next(iter(x.values()))
            out.append(float(v))
        else:
            out.append(float(x))
    return out


def _hour_to_time_label(h: int) -> str:
    """Map 0..23 to 12:00 AM .. 11:00 PM."""
    if h == 0:
        return "12:00 AM"
    if h < 12:
        return f"{h}:00 AM"
    if h == 12:
        return "12:00 PM"
    return f"{h - 12}:00 PM"


def _period_for_hour(h: int) -> str:
    if 0 <= h <= 5:
        return "Night (12 AM – 5 AM)"
    if 6 <= h <= 11:
        return "Morning (6 AM – 11 AM)"
    if 12 <= h <= 17:
        return "Afternoon (12 PM – 5 PM)"
    return "Evening (6 PM – 11 PM)"


def _build_workbook(hours: range, preds: list[float]) -> bytes:
    wb = Workbook()
    ws_counts = wb.active
    ws_counts.title = "Counts"

    headers = ["Time", "Hour (24h)", "Predicted count"]
    ws_counts.append(headers)
    for h, p in zip(hours, preds):
        ws_counts.append([_hour_to_time_label(h), h, round(p, 2)])

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, 4):
        cell = ws_counts.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in range(2, 26):
        for col in range(1, 4):
            c = ws_counts.cell(row=row, column=col)
            c.border = border
            c.alignment = Alignment(horizontal="left" if col <= 2 else "right", vertical="center")
        if row % 2 == 0:
            for col in range(1, 4):
                ws_counts.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F5F7FA")
    for col, width in enumerate((14, 12, 16), start=1):
        ws_counts.column_dimensions[get_column_letter(col)].width = width

    # Pivot-style summary by part of day
    df = pd.DataFrame({"hour": list(range(24)), "pred": preds})
    df["_period"] = df["hour"].map(_period_for_hour)
    summary = df.groupby("_period", sort=False).agg(sum_pred=("pred", "sum"), mean_pred=("pred", "mean")).reset_index()
    period_order = [
        "Night (12 AM – 5 AM)",
        "Morning (6 AM – 11 AM)",
        "Afternoon (12 PM – 5 PM)",
        "Evening (6 PM – 11 PM)",
    ]
    summary["_sort"] = summary["_period"].map({p: i for i, p in enumerate(period_order)})
    summary = summary.sort_values("_sort").drop(columns="_sort")

    ws_chart = wb.create_sheet("Chart_and_pivot")
    ws_chart.merge_cells("A1:D1")
    title = ws_chart["A1"]
    title.value = "Summary by time of day (pivot-style)"
    title.font = Font(size=14, bold=True, color="1E3A5F")
    title.alignment = Alignment(horizontal="left", vertical="center")

    piv_headers = ["Period of day", "Sum of predicted count", "Average per hour", "Hours in bucket"]
    ws_chart.append([])
    r = 3
    for j, htxt in enumerate(piv_headers, start=1):
        cell = ws_chart.cell(row=r, column=j)
        cell.value = htxt
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1
    hours_per: dict[str, int] = {}
    for h in range(24):
        p = _period_for_hour(h)
        hours_per[p] = hours_per.get(p, 0) + 1
    for _, prow in summary.iterrows():
        period_name = prow["_period"]
        s = float(prow["sum_pred"])
        m = float(prow["mean_pred"])
        n = hours_per.get(period_name, 6)
        ws_chart.append([period_name, round(s, 1), round(m, 2), n])
        crow = ws_chart.max_row
        for col in range(1, 5):
            c = ws_chart.cell(row=crow, column=col)
            c.border = border
            if col > 1:
                c.alignment = Alignment(horizontal="right")
    for col, width in enumerate((28, 18, 18, 14), start=1):
        ws_chart.column_dimensions[get_column_letter(col)].width = width

    chart_row = ws_chart.max_row + 3
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Predicted demand by time of day"
    chart.y_axis.title = "Predicted count"
    chart.x_axis.title = "Clock time"
    chart.height = 14
    chart.width = 22
    data_ref = Reference(ws_counts, min_col=3, min_row=1, max_row=25)
    cats = Reference(ws_counts, min_col=1, min_row=2, max_row=25)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.x_axis.tickLblSkip = 2
    chart.legend = None
    ws_chart.add_chart(chart, f"A{chart_row}")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _html_summary(target: date, preds: list[float]) -> str:
    peak_h = max(range(24), key=lambda i: preds[i])
    total = sum(preds)
    peak_label = _hour_to_time_label(peak_h)
    rows = "".join(
        f"<tr style='background-color:{'#f5f7fa' if h % 2 else '#ffffff'};'>"
        f"<td style='padding:8px 12px;border:1px solid #e2e8f0;'>{_hour_to_time_label(h)}</td>"
        f"<td style='padding:8px 12px;border:1px solid #e2e8f0;text-align:right;'>{preds[h]:.1f}</td>"
        f"<td style='padding:8px 12px;border:1px solid #e2e8f0;text-align:center;color:#64748b;'>{h}</td></tr>"
        for h in range(24)
    )
    fmt_date = target.strftime("%A, %B %d, %Y")
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width"/></head>
<body style="margin:0;padding:0;background-color:#edf2f7;font-family:Georgia,'Segoe UI',Arial,sans-serif;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background-color:#edf2f7;">
<tr><td align="center" style="padding:28px 12px;">
<table role="presentation" width="640" cellspacing="0" cellpadding="0" style="max-width:640px;background:#ffffff;border-radius:10px;overflow:hidden;box-shadow:0 4px 14px rgba(30,58,95,0.12);">
<tr><td style="background:linear-gradient(135deg,#1e3a5f 0%,#2d5a87 100%);padding:26px 28px;color:#ffffff;">
<p style="margin:0;font-size:13px;letter-spacing:0.12em;text-transform:uppercase;opacity:0.9;">Bike demand forecast</p>
<h1 style="margin:10px 0 0 0;font-size:24px;font-weight:600;line-height:1.3;">{fmt_date}</h1>
<p style="margin:12px 0 0 0;font-size:14px;opacity:0.95;line-height:1.5;">Scenario uses median historical weather by hour for each clock time. See the attached Excel workbook: <strong>Counts</strong> (full table) and <strong>Chart_and_pivot</strong> (summary + bar chart).</p>
</td></tr>
<tr><td style="padding:24px 28px 8px 28px;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:separate;border-spacing:12px 0;">
<tr>
<td style="width:33%;background:#f0f9ff;border-radius:8px;padding:16px;text-align:center;border:1px solid #bae6fd;">
<p style="margin:0;font-size:12px;color:#0369a1;text-transform:uppercase;">Peak time</p>
<p style="margin:8px 0 0 0;font-size:18px;font-weight:bold;color:#0c4a6e;">{peak_label}</p>
<p style="margin:4px 0 0 0;font-size:13px;color:#475569;">~{preds[peak_h]:.0f} rides</p></td>
<td style="width:33%;background:#f0fdf4;border-radius:8px;padding:16px;text-align:center;border:1px solid #bbf7d0;">
<p style="margin:0;font-size:12px;color:#15803d;text-transform:uppercase;">Sum (hourly)</p>
<p style="margin:8px 0 0 0;font-size:18px;font-weight:bold;color:#14532d;">{total:.0f}</p>
<p style="margin:4px 0 0 0;font-size:12px;color:#64748b;">Illustrative total</p></td>
<td style="width:33%;background:#faf5e4;border-radius:8px;padding:16px;text-align:center;border:1px solid #fde68a;">
<p style="margin:0;font-size:12px;color:#a16207;text-transform:uppercase;">Forecast date</p>
<p style="margin:8px 0 0 0;font-size:16px;font-weight:bold;color:#713f12;">{target.isoformat()}</p>
</td></tr></table>
</td></tr>
<tr><td style="padding:8px 28px 28px 28px;">
<p style="margin:0 0 12px 0;font-size:15px;font-weight:600;color:#1e293b;">Hourly predicted demand</p>
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;font-size:14px;">
<tr style="background-color:#1e3a5f;color:#ffffff;">
<th style="padding:10px 12px;text-align:left;border:1px solid #1e3a5f;">Time</th>
<th style="padding:10px 12px;text-align:right;border:1px solid #1e3a5f;">Predicted demand</th>
<th style="padding:10px 12px;text-align:center;border:1px solid #1e3a5f;">Hour (24h)</th>
</tr>
{rows}
</table>
</td></tr>
<tr><td style="padding:0 28px 24px 28px;font-size:12px;color:#64748b;line-height:1.5;">
<p style="margin:0;">Values are model outputs for the scenario described in the project docs. Open the .xlsx attachment for sortable data and the chart.</p>
</td></tr>
</table>
</td></tr></table>
</body></html>"""


def _send_email(subject: str, html_body: str, attachment: bytes, filename: str) -> None:
    src = os.environ["REPORT_EMAIL_FROM"]
    dst = os.environ["REPORT_EMAIL_TO"]
    region = os.environ.get("AWS_REGION", os.environ.get("AWS_DEFAULT_REGION", "us-east-1"))
    ses = boto3.client("ses", region_name=region)

    if not attachment:
        raise ValueError("Email attachment is empty; refusing to send without .xlsx")

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = src
    msg["To"] = dst

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html", "utf-8"))
    msg.attach(alt)

    # Explicit Excel MIME type; MIMEApplication base64-encodes by default. Use policy.SMTP for CRLF on send.
    att = MIMEApplication(
        attachment,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    att.add_header("Content-Disposition", "attachment", filename=filename)
    att.add_header("Content-Description", filename)
    msg.attach(att)

    raw = msg.as_bytes(policy=policy.SMTP)
    logger.info("SES SendRawEmail attachment %s (%d bytes), raw message %d bytes", filename, len(attachment), len(raw))
    ses.send_raw_email(Source=src, Destinations=[dst], RawMessage={"Data": raw})


def handler(event, context):
    endpoint = os.environ["ENDPOINT_NAME"]
    region = os.environ.get("AWS_REGION", os.environ.get("AWS_DEFAULT_REGION", "us-east-1"))
    tz_name = os.environ.get("REPORT_TIMEZONE", "America/New_York")
    tz = ZoneInfo(tz_name)

    override = os.environ.get("TARGET_DATE")
    if override:
        target = date.fromisoformat(override)
    else:
        target = (datetime.now(tz) + timedelta(days=1)).date()

    profile = _load_profile()
    instances = build_instances(target, profile)
    body = json.dumps({"instances": instances})

    sm = boto3.client("sagemaker-runtime", region_name=region)
    resp = sm.invoke_endpoint(
        EndpointName=endpoint,
        ContentType="application/json",
        Accept="application/json",
        Body=body.encode("utf-8"),
    )
    raw = resp["Body"].read().decode("utf-8")
    preds = _parse_predictions(raw)
    if len(preds) != 24:
        raise RuntimeError(f"Expected 24 predictions, got {len(preds)}")

    xlsx = _build_workbook(range(24), preds)
    fname = f"bike_demand_forecast_{target.isoformat()}.xlsx"
    subject = f"Bike demand forecast {target.isoformat()} (hourly)"
    html = _html_summary(target, preds)
    _send_email(subject, html, xlsx, fname)

    logger.info("Sent report for %s to %s", target, os.environ["REPORT_EMAIL_TO"])
    return {"statusCode": 200, "body": json.dumps({"forecast_date": target.isoformat(), "predictions": preds})}
